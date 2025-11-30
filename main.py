import openpyxl
from openpyxl.styles import Border, Side, Font, PatternFill

# Функция для создания Excel файла с диапазоном чисел в столбце A
def create_excel_file(filename, start, end):
    wb = openpyxl.Workbook()
    ws = wb.active
    for num in range(start, end + 1):
        ws.append([num])
    wb.save(filename)

# Создаем три файла с указанными диапазонами (на основе описания: 1-111, 2222-3333, и третий аналогичный)
create_excel_file('file1.xlsx', 1, 111)
create_excel_file('file2.xlsx', 2222, 3333)
create_excel_file('file3.xlsx', 1, 111)  # Если третий файл не specified, используем похожий; скорректируй, если нужно

# Функция для чтения данных из Excel файла (данные в столбце A)
def read_excel_data(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row[0])
    return data

# Читаем данные из трех файлов
data1 = read_excel_data('file1.xlsx')
data2 = read_excel_data('file2.xlsx')
data3 = read_excel_data('file3.xlsx')

# Объединяем в один список (матрицу)
matrix = data1 + data2 + data3

# Сортируем в порядке убывания
matrix = sorted(matrix, reverse=True)

# Создаем новый Excel файл
wb = openpyxl.Workbook()
ws = wb.active

# Записываем отсортированные данные в столбец A
for num in matrix:
    ws.append([num])

# Применяем стили: границы, шрифт, заливка (на основе "приоритет и очереди" — предполагаю форматирование для выделения)
thin_border = Side(style='thin', color='000000')
border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Желтая заливка для видимости

font = Font(bold=True, color='000000')  # Жирный шрифт

for row in ws.iter_rows(min_row=1, max_row=len(matrix), min_col=1, max_col=1):
    for cell in row:
        cell.border = border
        cell.fill = fill
        cell.font = font

# Сохраняем файл
wb.save('output.xlsx')

print("Файл 'output.xlsx' создан с отсортированной матрицей и стилями.")
